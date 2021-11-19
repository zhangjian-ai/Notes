import socket
import openpyxl as openpyxl

# 创建客户端socket对象
client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
# 服务端IP地址和端口号元组
server_address = ('127.0.0.1', 8222)
# 客户端连接指定的IP地址和端口号
client.connect(server_address)

# 导出数据标题
title = ['cpu_used_rate', 'mem_used_rate', 'mem_used', 'mem_free', 'mem_available', 'receive', 'send']

# 初始行数
row = 1

# 创建sheet，并指定为第一个sheet。创建表格对象时会默认创建一个sheet
excel = openpyxl.Workbook()
sheet = excel.create_sheet('ExportData', 0)

while row <= 20:
    # 保存数据
    for col, value in enumerate(title):
        sheet.cell(row, col + 1).value = value

    # 客户端发送数据
    client.send('next'.encode())
    # 客户端接收数据
    server_data = client.recv(1024).decode('utf-8')
    print("server >> ", server_data)

    title = server_data.split('-')

    row += 1


# 保存数据
excel.save('server.xlsx')

client.send('exit'.encode('utf-8'))
client.close()
