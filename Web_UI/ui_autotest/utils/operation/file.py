import configparser
import json

import openpyxl as openpyxl
import ruamel.yaml
import yaml
import jsonpath


def load_json(filepath):
    """
    读取json文件
    :param filepath:
    :return: dict
    """
    with open(filepath, 'r', encoding='utf8') as fp:
        data = json.load(fp)
    return data


def dump_json(filepath, content):
    """
    写入json文件
    :param filepath:
    :param content:
    :return:
    """
    with open(filepath, 'w', encoding='utf8') as fp:
        # 将ensure_ascii 设置为false，避免将汉字转成ascii码
        # indent 参数让json文本保持缩进格式，数字表述缩进字符数
        json.dump(content, fp, ensure_ascii=False, indent=2)


def load_yaml(filepath):
    """
    读取yaml文件
    :param filepath:
    :return: dict
    """
    with open(filepath, 'rb') as fp:
        data = yaml.safe_load(fp)
    return data


def dump_yaml(filepath, content):
    """
    写入yaml文件
    :param filepath:
    :param content:
    :return:
    """
    with open(filepath, 'w', encoding='utf8') as fp:
        # allow_unicode参数为true时，才会进行编码。否则写入的文件都是进过url编码的
        # default_flow_style 表示dump后的字典数据全部以yml格式显示,默认为为True
        # sort_keys 不给key排序，这样能保持原来字典的顺序写入。为True时按字母的排序展示，默认为为True
        # indent 参数让yaml文本保持缩进格式。有下次，列表符号不会跟随缩进
        # yaml.safe_dump(content, fp, allow_unicode=True, default_flow_style=False, sort_keys=False, indent=4)

        # ----为了dump保持缩进，这里采用另一个模块实现
        r_yaml = ruamel.yaml.YAML()
        r_yaml.indent(sequence=4, offset=2)
        r_yaml.dump(content, fp)


def load_ini(filepath):
    """
    固定读取conf文件夹里面的配置文件
    :return: dict
    """
    loader = configparser.ConfigParser()
    loader.read(filepath, encoding='utf8')

    ini_dict = {}
    for section in loader.sections():
        section_dict = {}
        for option in loader.options(section):
            section_dict.setdefault(option, loader.get(section, option))
        ini_dict.setdefault(section, section_dict)

    return ini_dict


def dump_ini(filepath, content: dict):
    """
    写入ini文件信息
    :param filepath:
    :param content:
    :return:
    """
    pass


def load_excel(path: str):
    """
    返回一个包含多个列表的列表：[[],[],[]...]
    """
    table = openpyxl.load_workbook(path)
    sheet_names = table.get_sheet_names()

    # 默认读取第一个sheet页
    sheet = table.get_sheet_by_name(sheet_names[0])
    rows = sheet.rows
    sheet_list = []

    for row in rows:
        row_list = []
        for cell in row:
            row_list.append(cell.value)
        sheet_list.append(row_list)

    return sheet_list


def dump_excel(path: str, data: list):
    """
    按行写入excel
    :param path:
    :param data: [[],[],[]...]
    :return:
    """
    excel = openpyxl.Workbook()

    # 创建sheet，并指定为第一个sheet。创建表格对象时会默认创建一个sheet
    sheet = excel.create_sheet('ExportData', 0)

    for row_no, row in enumerate(data):
        for col_no, value in enumerate(row):
            sheet.cell(row_no + 1, col_no + 1).value = value

    excel.save(path)


def get_case_id(filepath, case_path):
    """
    获取用例下面的多个caseId
    :param filepath:
    :param case_path:
    :return: list => [id, id, id]
    """
    case_id = jsonpath.jsonpath(load_yaml(filepath), f"$.cases..[?(@.CasePath == '{case_path}')]..CaseID")
    return case_id


def get_case_info(filepath, case_id):
    """
    获取用例ID对应的用例描述信息
    :param filepath:
    :param case_id:
    :return: dict
    """
    case_info = jsonpath.jsonpath(load_yaml(filepath), f"$.datas..[?(@.CaseID == '{case_id}')]")
    return case_info[0] if case_info else {}
